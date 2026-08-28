"""Convert regional Fuel Crisis Policy tracker workbooks into standalone HTML
addons consumed by ``small_dashboard_integrated_w_policy``.

Pipeline:
    data/text/policy_tracker/<region>.xlsx
      -> src/text/plotting/addons/<tracker>/<region>_policy_addon.html

CLI:
    poetry run po text build-policy-addons --region eap
    python -m text.plotting.policy_dashboards --only eap sar

The HTML files are self-contained (embedded JSON, no external JS) so they
survive corporate sanitizers when embedded as iframe ``srcdoc`` inside the
host dashboard.
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import re
import sys
import zipfile
from collections import Counter
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook

from text.plotting.trackers import (
    DEFAULT_TRACKER,
    TRACKERS,
    addon_filename,
    get_tracker,
    tracker_dir,
    tracker_label,
    workbook_dir,
)


PROJECT_ROOT = Path(__file__).resolve().parents[3]
DEFAULT_INPUT_DIR = PROJECT_ROOT / "data" / "text" / "policy_tracker"
DEFAULT_OUTPUT_DIR = Path(__file__).resolve().parent / "addons"


DEFAULT_LABEL_ORDER = [
    "Communication",
    "General consumption smoothing",
    "Guaranteeing essential services",
    "Reduce demand - higher prices",
    "Reduce demand - restricting quantities",
    "Secure supply",
    "Support to business",
    "Support to households",
    "No confirmed measure",
    "Proposed / structural",
    "Shield consumers",
    "Pass-through pricing",
]

DEFAULT_COLORS = {
    "Communication": "#4f81bd",
    "General consumption smoothing": "#c0504d",
    "Guaranteeing essential services": "#9bbb59",
    "Reduce demand - higher prices": "#7030a0",
    "Reduce demand - restricting quantities": "#8064a2",
    "Reduce demand": "#8064a2",  # legacy input; normalized before output
    "Secure supply": "#4bacc6",
    "Support to business": "#f79646",
    "Support to households": "#ffc000",
    "No confirmed measure": "#a5a5a5",
    "Proposed / structural": "#7f6000",
    "Shield consumers": "#ffd966",
    "Pass-through pricing": "#7030a0",
}

EXTRA_PALETTE = [
    "#70ad47",
    "#5b9bd5",
    "#ed7d31",
    "#a5a5a5",
    "#4472c4",
    "#843c0c",
    "#2f75b5",
    "#548235",
    "#7030a0",
    "#bf9000",
]

FIELD_ALIASES = {
    "country": [
        "Country",
        "Economy",
        "Country/economy",
        "Country / economy",
        "Country or economy",
        "Jurisdiction",
        "Market",
        "Member State",
        "Member state",
        "Location",
        "Geography",
    ],
    "number": [
        "#",
        "No",
        "No.",
        "ID",
        "Policy ID",
        "Measure ID",
        "Tracker ID",
        "Row ID",
        "Policy #",
    ],
    "policy": [
        "Policy",
        "Policy measure",
        "Policy Measure",
        "Measure",
        "Policy/measure",
        "Policy / measure",
        "Policy title",
        "Measure title",
        "Instrument",
        "Policy instrument",
        "Response measure",
        "Policy response",
        "Intervention",
        "Action",
        "Policy name",
        "Measure / instrument",
    ],
    "description": [
        "Policy Description",
        "Description",
        "Policy description",
        "Measure description",
        "Details",
        "Description / details",
        "Summary",
        "Narrative",
        "Notes",
        "Policy details",
        "Detail",
        "Evidence summary",
        "Policy evidence",
        "Measure details",
    ],
    "label": [
        "Label",
        "Policy label",
        "Policy Label",
        "Dashboard label",
        "Dashboard Label",
        "Type",
        "Measure type",
        "Policy type",
        "Instrument type",
        "Response category",
        "Broad category",
        "Classification",
    ],
    "category": [
        "Category",
        "Policy category",
        "Policy Category",
        "v6_category",
        "V6 category",
    ],
    "subcategory": [
        "Subcategory",
        "Sub-category",
        "Sub category",
        "Policy subcategory",
        "Policy Subcategory",
        "v6_subcategory",
        "V6 subcategory",
    ],
    "date_status": [
        "Active or Proposed Date",
        "Active/Proposed Date",
        "Active or proposed date",
        "Status/date",
        "Status / date",
        "Status Date",
        "Implementation status/date",
        "Implementation status / date",
        "Date",
        "Timing",
        "Effective date",
        "Active date",
        "Proposed date",
        "Period",
        "Current / ongoing",
    ],
    "status": [
        "Status",
        "Implementation status",
        "Current status",
        "Policy status",
        "Active/proposed",
        "Active or proposed",
    ],
    "source": [
        "Source",
        "Sources",
        "Source name",
        "Main source",
        "Evidence source",
        "Reference",
        "References",
        "Primary source",
        "Source title",
        "Source(s)",
    ],
    "source_url": [
        "Source URL",
        "Source URLs",
        "Source url",
        "URL",
        "Urls",
        "Links",
        "Link",
        "Source link",
        "Primary URL",
        "Evidence URL",
        "Citation URL",
        "Source URL(s)",
    ],
    "evaluation": [
        "Evaluation",
        "Verification",
        "Verification status",
        "Evidence grade",
        "Confidence",
        "Source quality",
        "Assessment",
        "Review status",
        "Validation",
        "Reliability",
    ],
    "reason": [
        "Reason",
        "Verification note",
        "Verification notes",
        "Notes on verification",
        "Rationale",
        "Comment",
        "Comments",
        "Review note",
        "Evidence note",
        "Source note",
        "Why included",
    ],
    "fuel_product": [
        "Fuel",
        "Fuels",
        "Product",
        "Products",
        "Fuel products",
        "Fuel/product",
        "Product(s)",
    ],
    "direction": [
        "Direction",
        "Action direction",
        "Policy direction",
        "Change",
        "Stance",
    ],
    "agency": [
        "Agency",
        "Ministry",
        "Implementing agency",
        "Authority",
        "Institution",
        "Responsible agency",
    ],
}

NON_POLICY_EXCLUSION_PHRASES = [
    "scope definition",
    "no verified current discretionary fuel relief",
    "no verified current fuel-price policy change",
    "no confirmed current discretionary fuel relief",
    "no confirmed current fuel-price policy change",
    "source gap",
    "coverage gap",
    "watchlist only",
    "monitoring lead only",
]

# World Bank Pacific Island Country members for the EAP "PICs only (12)" view.
WB_PIC_MEMBERS = [
    "Fiji",
    "Kiribati",
    "RMI",
    "FSM",
    "Nauru",
    "Palau",
    "PNG",
    "Samoa",
    "Solomon Islands",
    "Tonga",
    "Tuvalu",
    "Vanuatu",
]

REGIONS: List[Dict[str, Any]] = [
    {"key": "ssa", "display_name": "Sub-Saharan Africa"},
    {
        "key": "eca",
        "display_name": "Europe and Central Asia",
        "exclude_if_any_field_contains": NON_POLICY_EXCLUSION_PHRASES,
    },
    {"key": "menaap", "display_name": "MENAAP"},
    {
        "key": "sar",
        "display_name": "South Asia",
        "exclude_countries": ["Afghanistan", "Pakistan"],
    },
    {"key": "lac", "display_name": "Latin America and Caribbean"},
    {
        "key": "eap",
        "display_name": "East Asia and Pacific",
        "world_bank_pic_members": WB_PIC_MEMBERS,
    },
]

# Workbooks share a unified schema: a `Policies` sheet alongside metadata
# sheets. The converter prefers the canonical name and falls back to the
# first sheet if a workbook hasn't been migrated.
CANONICAL_SHEET = "Policies"


def norm_key(value: Any) -> str:
    text = "" if value is None else str(value)
    text = text.strip().lower()
    text = re.sub(r"[\s\n\r\t]+", " ", text)
    text = text.replace("&", "and")
    text = re.sub(r"[^a-z0-9]+", "", text)
    return text


def norm_country(value: Any) -> str:
    return re.sub(r"\s+", " ", "" if value is None else str(value).strip()).casefold()


def as_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return f"{value:g}"
    return re.sub(r"\s+", " ", str(value)).strip()


def clean_text(text: str) -> str:
    return re.sub(r"\s+", " ", text or "").strip()


def find_value(row_by_norm_header: Dict[str, str], field: str) -> str:
    for alias in FIELD_ALIASES[field]:
        key = norm_key(alias)
        if key in row_by_norm_header:
            value = clean_text(row_by_norm_header.get(key, ""))
            if value:
                return value
    return ""


def detect_header_row(ws, max_scan_rows: int = 15) -> Tuple[int, List[str]]:
    country_keys = {norm_key(x) for x in FIELD_ALIASES["country"]}
    policy_keys = {norm_key(x) for x in FIELD_ALIASES["policy"]}
    label_keys = {norm_key(x) for x in FIELD_ALIASES["label"]}
    desc_keys = {norm_key(x) for x in FIELD_ALIASES["description"]}

    best_score = -1
    best_row = 1
    best_headers: List[str] = []

    scan_limit = min(max_scan_rows, ws.max_row or max_scan_rows)
    for row_idx in range(1, scan_limit + 1):
        headers = [as_text(c.value) for c in ws[row_idx]]
        keys = {norm_key(h) for h in headers if h}
        score = 0
        score += 4 if keys & country_keys else 0
        score += 4 if keys & policy_keys else 0
        score += 3 if keys & label_keys else 0
        score += 2 if keys & desc_keys else 0
        score += min(4, sum(1 for h in headers if h))
        if score > best_score:
            best_score = score
            best_row = row_idx
            best_headers = headers

    if best_score < 8:
        raise ValueError(
            f"Could not confidently detect a header row in sheet '{ws.title}'. "
            "Check that the sheet has country and policy/measure columns."
        )
    return best_row, best_headers


PRICE_DEMAND_KEYWORDS = [
    "price increase",
    "prices increase",
    "price adjustment",
    "higher price",
    "raise price",
    "tariff increase",
    "rate increase",
    "excise increase",
    "tax increase",
    "levy increase",
    "subsidy reduction",
    "reduced subsidy",
    "phase-down",
    "phase down",
    "pass-through",
    "pass through",
    "cost-reflective",
    "market price",
    "wholesale price increases",
]


def split_reduce_demand_label(
    label: str, policy: str, description: str, reason: str = ""
) -> str:
    """Map legacy ``Reduce demand`` / ``Pass-through pricing`` into the revised two-part typology."""
    label = clean_text(label)
    text = f"{policy} {description} {reason}".lower()
    if label == "Pass-through pricing":
        return "Reduce demand - higher prices"
    if label != "Reduce demand":
        return label
    if any(k in text for k in PRICE_DEMAND_KEYWORDS):
        return "Reduce demand - higher prices"
    return "Reduce demand - restricting quantities"


def infer_label(policy: str, description: str) -> str:
    text = f"{policy} {description}".lower()
    if any(
        x in text
        for x in [
            "work from home",
            "wfh",
            "odd-even",
            "ration",
            "reduce travel",
            "carpool",
            "conservation",
            "energy saving",
            "demand",
        ]
    ):
        return "Reduce demand - restricting quantities"
    if any(
        x in text
        for x in [
            "strategic reserve",
            "reserve",
            "import",
            "supply",
            "stockholding",
            "storage",
            "procurement",
            "cargo",
            "refinery",
            "secure",
        ]
    ):
        return "Secure supply"
    if any(
        x in text
        for x in [
            "cash",
            "welfare",
            "household",
            "voucher",
            "low-income",
            "low income",
            "family",
            "subsidy to households",
        ]
    ):
        return "Support to households"
    if any(
        x in text
        for x in [
            "business",
            "firm",
            "msme",
            "transport operator",
            "farmer",
            "fisher",
            "loan",
            "guarantee",
        ]
    ):
        return "Support to business"
    if any(
        x in text
        for x in [
            "emergency declaration",
            "essential",
            "hospital",
            "school",
            "electricity",
            "public transport",
            "bus service",
        ]
    ):
        return "Guaranteeing essential services"
    if any(
        x in text
        for x in [
            "tax",
            "excise",
            "vat",
            "price cap",
            "price freeze",
            "stabilization",
            "stabilisation",
            "levy",
            "tariff",
            "discount",
        ]
    ):
        return "General consumption smoothing"
    if any(
        x in text
        for x in [
            "monitor",
            "communication",
            "appeal",
            "campaign",
            "warning",
            "task force",
            "committee",
            "public",
            "briefing",
        ]
    ):
        return "Communication"
    return "Communication"


def normalize_status_date(
    status_date: str, status: str, policy: str, description: str
) -> str:
    status_date = clean_text(status_date)
    status = clean_text(status)
    if status_date:
        return status_date
    combined = f"{status} {policy} {description}".lower()
    if status:
        if status.lower().startswith(
            ("active", "implemented", "current", "ongoing", "in force")
        ):
            return f"Active {status}"
        if status.lower().startswith(
            ("proposed", "planned", "under consideration", "draft", "pending")
        ):
            return f"Proposed {status}"
        return status
    if any(
        x in combined
        for x in ["proposed", "planned", "under consideration", "draft", "pending"]
    ):
        return "Proposed"
    if any(
        x in combined
        for x in ["reviewed through", "no confirmed", "not located", "no verified"]
    ):
        return "Reviewed"
    return "Active"


def compact_join(parts: Iterable[Tuple[str, str]]) -> str:
    out = []
    seen = set()
    for label, value in parts:
        value = clean_text(value)
        if not value:
            continue
        key = (label.lower(), value.lower())
        if key in seen:
            continue
        seen.add(key)
        out.append(f"{label}: {value}" if label else value)
    return " ".join(out)


def normalize_row(raw: Dict[str, str], seq_num: int) -> Optional[Dict[str, str]]:
    country = find_value(raw, "country")
    policy = find_value(raw, "policy")

    direction = find_value(raw, "direction")
    if policy and direction and direction.lower() not in policy.lower():
        policy = f"{policy} — {direction}"

    description = find_value(raw, "description")
    if not description:
        description = compact_join(
            [
                ("Fuels", find_value(raw, "fuel_product")),
                ("Direction", direction),
                ("Agency", find_value(raw, "agency")),
                ("Status", find_value(raw, "status")),
                ("Verification note", find_value(raw, "reason")),
            ]
        )

    details_to_append = compact_join(
        [
            ("Fuels", find_value(raw, "fuel_product")),
            ("Direction", direction),
        ]
    )
    if details_to_append and details_to_append.lower() not in description.lower():
        description = clean_text(f"{description} {details_to_append}")

    if not country and not policy and not description:
        return None
    if not country or not policy:
        return None

    label = find_value(raw, "label") or infer_label(policy, description)
    label = split_reduce_demand_label(
        label, policy, description, find_value(raw, "reason")
    )
    status = find_value(raw, "status")
    status_date = normalize_status_date(
        find_value(raw, "date_status"), status, policy, description
    )
    number = find_value(raw, "number") or str(seq_num)
    source = find_value(raw, "source")
    source_url = find_value(raw, "source_url")
    evaluation = find_value(raw, "evaluation")
    reason = find_value(raw, "reason")

    return {
        "Country": country,
        "#": number,
        "Policy": policy,
        "Policy Description": description,
        "Label": label,
        "Category": find_value(raw, "category"),
        "Subcategory": find_value(raw, "subcategory"),
        "Active or Proposed Date": status_date,
        "Source": source,
        "Source URL": source_url,
        "Evaluation": evaluation,
        "Reason": reason,
    }


def load_policy_rows(
    xlsx_path: Path, sheet_name: Optional[str]
) -> Tuple[List[Dict[str, str]], str, int]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        elif sheet_name:
            print(
                f"  sheet '{sheet_name}' not found in {xlsx_path.name}; falling back to '{wb.sheetnames[0]}'"
            )
            ws = wb[wb.sheetnames[0]]
        else:
            ws = wb[wb.sheetnames[0]]

        header_row, headers = detect_header_row(ws)
        norm_headers = [norm_key(h) for h in headers]

        rows: List[Dict[str, str]] = []
        seq = 1
        for row_values in ws.iter_rows(min_row=header_row + 1, values_only=True):
            raw: Dict[str, str] = {}
            for idx, value in enumerate(row_values):
                if idx >= len(norm_headers):
                    continue
                key = norm_headers[idx]
                if not key:
                    continue
                raw[key] = as_text(value)
            normalized = normalize_row(raw, seq)
            if normalized:
                rows.append(normalized)
                seq += 1
        return rows, ws.title, header_row
    finally:
        wb.close()


def apply_region_filters(
    rows: List[Dict[str, str]], region_cfg: Dict[str, Any]
) -> Tuple[List[Dict[str, str]], int]:
    exclude_countries = {
        norm_country(c) for c in region_cfg.get("exclude_countries", [])
    }
    exclude_labels = {
        clean_text(x).casefold() for x in region_cfg.get("exclude_labels", [])
    }
    exclude_phrases = [
        x.casefold() for x in region_cfg.get("exclude_if_any_field_contains", [])
    ]

    kept = []
    excluded = 0
    for row in rows:
        if norm_country(row.get("Country", "")) in exclude_countries:
            excluded += 1
            continue
        if clean_text(row.get("Label", "")).casefold() in exclude_labels:
            excluded += 1
            continue
        haystack = " | ".join(str(v) for v in row.values()).casefold()
        if exclude_phrases and any(phrase in haystack for phrase in exclude_phrases):
            excluded += 1
            continue
        kept.append(row)
    return kept, excluded


def sort_countries_by_count(rows: List[Dict[str, str]]) -> List[str]:
    counts = Counter(r["Country"] for r in rows)
    return sorted(counts.keys(), key=lambda c: (-counts[c], c.casefold()))


def build_country_groups(
    rows: List[Dict[str, str]], region_cfg: Dict[str, Any]
) -> Dict[str, List[str]]:
    all_countries = sort_countries_by_count(rows)
    groups: Dict[str, List[str]] = {"All countries and economies": all_countries}

    active = [
        c
        for c in all_countries
        if any(
            r["Country"] == c
            and r["Active or Proposed Date"].lower().startswith("active")
            for r in rows
        )
    ]
    proposed = [
        c
        for c in all_countries
        if any(
            r["Country"] == c
            and r["Active or Proposed Date"].lower().startswith("proposed")
            for r in rows
        )
    ]
    regional = [
        c
        for c in all_countries
        if re.search(
            r"\b(all|regional|region|european union|eu|gulf cooperation|gcc|caricom|oecs|pacific|multi-country)\b",
            c,
            flags=re.I,
        )
    ]

    if active and len(active) != len(all_countries):
        groups["Countries/economies with active rows"] = active
    if proposed:
        groups["Countries/economies with proposed rows"] = proposed
    if regional:
        groups["Regional / multi-country groups"] = regional

    for name, members in region_cfg.get("country_groups", {}).items():
        clean_members = [m for m in members if m in all_countries]
        if clean_members:
            groups[name] = clean_members

    wb_pics = region_cfg.get("world_bank_pic_members")
    if wb_pics:
        pic_members = [m for m in wb_pics if m in all_countries]
        if pic_members:
            groups["World Bank PICs only (12)"] = pic_members

    return groups


def build_display_names(countries: Iterable[str]) -> Dict[str, str]:
    overrides = {
        "Solomon Islands": "Solomon\nIslands",
        "Hong Kong SAR, China": "Hong Kong SAR,\nChina",
        "Macao SAR, China": "Macao SAR,\nChina",
        "Taiwan, China": "Taiwan,\nChina",
        "Brunei Darussalam": "Brunei\nDarussalam",
        "Korea, Rep.": "Korea,\nRep.",
        "Timor Leste": "Timor\nLeste",
        "New Caledonia": "New\nCaledonia",
        "French Polynesia": "French\nPolynesia",
        "Dominican Republic": "Dominican\nRepublic",
        "Trinidad and Tobago": "Trinidad and\nTobago",
        "Bosnia and Herzegovina": "Bosnia and\nHerzegovina",
        "North Macedonia": "North\nMacedonia",
        "Russian Federation": "Russian\nFederation",
        "United Arab Emirates": "United Arab\nEmirates",
        "Saudi Arabia": "Saudi\nArabia",
        "South Africa": "South\nAfrica",
        "Cote d'Ivoire": "Cote\nd'Ivoire",
        "Côte d’Ivoire": "Côte\nd’Ivoire",
        "Congo, Dem. Rep.": "Congo,\nDem. Rep.",
        "Congo, Rep.": "Congo,\nRep.",
        "All (Pacific)": "All\n(Pacific)",
    }
    out = {}
    for country in countries:
        if country in overrides:
            out[country] = overrides[country]
        elif len(country) > 14 and " " in country:
            parts = country.split()
            mid = max(1, len(parts) // 2)
            out[country] = " ".join(parts[:mid]) + "\n" + " ".join(parts[mid:])
    return out


def labels_for_rows(
    rows: List[Dict[str, str]], region_cfg: Dict[str, Any]
) -> List[str]:
    present = {clean_text(r.get("Label", "")) or "Unclassified" for r in rows}
    configured_order = region_cfg.get("label_order") or DEFAULT_LABEL_ORDER
    labels = [label for label in configured_order if label in present]
    labels += sorted(
        [label for label in present if label not in set(labels)], key=str.casefold
    )
    return labels


def colors_for_labels(labels: List[str], region_cfg: Dict[str, Any]) -> Dict[str, str]:
    colors = dict(DEFAULT_COLORS)
    colors.update(region_cfg.get("colors", {}))
    extra_idx = 0
    for label in labels:
        if label not in colors:
            colors[label] = EXTRA_PALETTE[extra_idx % len(EXTRA_PALETTE)]
            extra_idx += 1
    return {label: colors[label] for label in labels}


def build_dashboard_data(
    rows: List[Dict[str, str]],
    region_cfg: Dict[str, Any],
    xlsx_path: Path,
    sheet_name: str,
    excluded_count: int,
) -> Tuple[Dict[str, Any], Dict[str, str]]:
    labels = labels_for_rows(rows, region_cfg)
    groups = build_country_groups(rows, region_cfg)
    countries = sort_countries_by_count(rows)
    metadata = {
        "generated_on": dt.date.today().isoformat(),
        "source_file": xlsx_path.name,
        "source_sheet": sheet_name,
        "row_count": len(rows) + excluded_count,
        "country_count": len({r["Country"] for r in rows}),
        "dashboard_version": "policy_dashboards_v1",
        "region": region_cfg.get("display_name", region_cfg.get("key", "Region")),
        "included_rows": len(rows),
    }
    if excluded_count:
        metadata["excluded_rows"] = excluded_count
    if region_cfg.get("exclude_countries"):
        metadata["excluded_countries"] = region_cfg.get("exclude_countries")

    data = {
        "policies": rows,
        "labels": labels,
        "countryGroups": groups,
        "displayName": build_display_names(countries),
        "metadata": metadata,
    }
    return data, colors_for_labels(labels, region_cfg)


def make_html(data: Dict[str, Any], colors: Dict[str, str], chart_title: str) -> str:
    data_json = json.dumps(data, ensure_ascii=False, indent=2)
    colors_json = json.dumps(colors, ensure_ascii=False, indent=2)
    safe_title = json.dumps(chart_title, ensure_ascii=False)

    return f"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8" />
<meta name="viewport" content="width=device-width, initial-scale=1" />
<title>Fuel Crisis Policy Dashboard</title>
<style>
  :root {{
    --blue: #4f81bd;
    --red: #c0504d;
    --green: #9bbb59;
    --purple: #8064a2;
    --teal: #4bacc6;
    --orange: #f79646;
    --gold: #ffc000;
    --grid: #d9d9d9;
    --text: #555;
    --muted: #777;
    --panel: #ffffff;
    --border: #d8d8d8;
  }}
  html, body {{
    margin: 0;
    padding: 0;
    background: #f5f6f8;
    font-family: Calibri, Arial, Helvetica, sans-serif;
    color: var(--text);
  }}
  .page {{ max-width: 1180px; margin: 24px auto; padding: 0 16px 32px; }}
  .dashboard {{ background: #fff; border: 1px solid #d0d0d0; box-shadow: 0 1px 4px rgba(0,0,0,0.08); padding: 20px 24px 28px; }}
  .controls {{ display: grid; grid-template-columns: minmax(220px, 1.2fr) minmax(160px, .8fr) minmax(160px, .8fr) minmax(180px, .8fr); gap: 12px; align-items: end; margin-bottom: 14px; }}
  .control-group label {{ display: block; font-size: 12px; color: #666; margin-bottom: 4px; text-transform: uppercase; letter-spacing: .03em; }}
  select, input[type="text"] {{ width: 100%; box-sizing: border-box; border: 1px solid #cfcfcf; border-radius: 4px; padding: 7px 8px; font-size: 14px; background: #fff; color: #333; }}
  .kpis {{ display: grid; grid-template-columns: repeat(4, minmax(0, 1fr)); gap: 10px; margin-bottom: 4px; }}
  .kpi {{ border: 1px solid #e2e2e2; border-radius: 6px; padding: 8px 10px; background: #fbfbfb; }}
  .kpi .value {{ font-size: 24px; line-height: 1.1; color: #444; }}
  .kpi .label {{ font-size: 12px; color: #777; }}
  .chart-title {{ text-align: center; font-size: 25px; font-weight: 400; margin: 18px 0 0; color: #555; }}
  .chart-subtitle {{ text-align: center; font-size: 12px; color: #888; margin: 4px 0 10px; }}
  .chart-wrap {{ position: relative; width: 100%; overflow-x: auto; padding-bottom: 4px; }}
  svg {{ display: block; margin: 0 auto; background: #fff; min-width: 900px; }}
  .axis text {{ font-size: 15px; fill: #666; }}
  .grid line {{ stroke: var(--grid); stroke-width: 1.2; }}
  .axis-line {{ stroke: #cfcfcf; stroke-width: 1.2; }}
  .bar-segment {{ cursor: help; shape-rendering: crispEdges; }}
  .bar-segment:hover {{ filter: brightness(0.94); stroke: #333; stroke-width: 1px; }}
  .legend {{ display: flex; flex-wrap: wrap; justify-content: center; align-items: center; gap: 12px 20px; max-width: 900px; margin: 8px auto 0; font-size: 17px; color: #666; }}
  .legend-item {{ display: flex; align-items: center; gap: 5px; white-space: nowrap; }}
  .legend-swatch {{ width: 10px; height: 10px; display: inline-block; }}
  .tooltip {{ position: fixed; z-index: 20; pointer-events: none; display: none; width: min(440px, calc(100vw - 32px)); background: rgba(255,255,255,0.98); color: #333; border: 1px solid #bdbdbd; border-radius: 6px; box-shadow: 0 4px 18px rgba(0,0,0,0.18); padding: 10px 12px; font-size: 13px; line-height: 1.35; }}
  .tooltip h3 {{ font-size: 15px; margin: 0 0 5px; color: #333; }}
  .tooltip .meta {{ margin: 0 0 7px; color: #666; font-size: 12px; }}
  .tooltip ul {{ margin: 6px 0 0 18px; padding: 0; max-height: 260px; overflow: auto; }}
  .tooltip li {{ margin: 0 0 6px; }}
  .tooltip strong {{ color: #333; }}
  .detail-panel {{ margin-top: 20px; display: grid; grid-template-columns: 1fr; gap: 10px; }}
  .detail-card {{ border: 1px solid #e2e2e2; border-radius: 6px; background: #fff; padding: 12px 14px; }}
  .detail-card h2 {{ font-size: 17px; font-weight: 600; margin: 0 0 8px; color: #444; }}
  .detail-card p {{ margin: 4px 0; font-size: 13px; color: #666; }}
  .policy-list {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(260px, 1fr)); gap: 8px; margin-top: 8px; }}
  .policy-card {{ border-left: 5px solid #ccc; background: #fafafa; padding: 8px 10px; min-height: 72px; }}
  .policy-card .policy-title {{ font-weight: 600; color: #333; font-size: 13px; }}
  .policy-card .policy-desc {{ color: #666; font-size: 12px; margin-top: 3px; }}
  .policy-card .policy-foot {{ color: #888; font-size: 11px; margin-top: 5px; }}
  .note {{ margin-top: 14px; font-size: 12px; color: #777; }}
  @media (max-width: 820px) {{ .controls {{ grid-template-columns: 1fr 1fr; }} .kpis {{ grid-template-columns: 1fr 1fr; }} }}
  @media (max-width: 540px) {{ .controls, .kpis {{ grid-template-columns: 1fr; }} .dashboard {{ padding: 14px; }} }}
</style>
</head>
<body>
<div class="page">
  <div class="dashboard">
    <div class="controls">
      <div class="control-group"><label for="groupSelect">Country view</label><select id="groupSelect"></select></div>
      <div class="control-group"><label for="categorySelect">Policy category</label><select id="categorySelect"></select></div>
      <div class="control-group"><label for="statusSelect">Status</label><select id="statusSelect"><option value="all">Active and proposed</option><option value="active">Active only</option><option value="proposed">Proposed only</option></select></div>
      <div class="control-group"><label for="titleInput">Chart title</label><input id="titleInput" type="text" value={safe_title} /></div>
    </div>
    <div class="kpis">
      <div class="kpi"><div class="value" id="kpiPolicies">0</div><div class="label">Policies in current view</div></div>
      <div class="kpi"><div class="value" id="kpiCountries">0</div><div class="label">Countries/economies shown</div></div>
      <div class="kpi"><div class="value" id="kpiLabels">0</div><div class="label">Policy labels present</div></div>
      <div class="kpi"><div class="value" id="kpiMax">0</div><div class="label">Largest country total</div></div>
    </div>
    <h1 id="chartTitle" class="chart-title"></h1>
    <div class="chart-subtitle" id="subtitle"></div>
    <div class="chart-wrap">
      <svg id="chart" aria-label="Stacked bar chart of fuel-crisis policy responses by country and policy label"></svg>
      <div id="tooltip" class="tooltip"></div>
    </div>
    <div id="legend" class="legend"></div>
    <div class="detail-panel"><div class="detail-card" id="detailCard"><h2>Policy details</h2><p>Hover over a shaded bar segment to see an info tip. Click a segment to pin its policy details here.</p></div></div>
    <div class="note">Note: the workbook may contain month-level text or review-status text rather than true daily effective dates. The title shows the dashboard's last update date and remains editable; the chart counts the rows currently selected by the filters. Countries/economies are sorted from left to right by the number of policies in the current filtered view.</div>
  </div>
</div>
<script>
const DASHBOARD_DATA = {data_json};
const COLORS = {colors_json};
const state = {{ pinned: null }};

const groupSelect = document.getElementById("groupSelect");
const categorySelect = document.getElementById("categorySelect");
const statusSelect = document.getElementById("statusSelect");
const titleInput = document.getElementById("titleInput");
const chartTitle = document.getElementById("chartTitle");
const subtitle = document.getElementById("subtitle");
const svg = document.getElementById("chart");
const tooltip = document.getElementById("tooltip");
const legend = document.getElementById("legend");
const detailCard = document.getElementById("detailCard");

function initControls() {{
  Object.keys(DASHBOARD_DATA.countryGroups).forEach((name, index) => {{
    const opt = document.createElement("option");
    opt.value = name;
    opt.textContent = name;
    if (index === 0) opt.selected = true;
    groupSelect.appendChild(opt);
  }});
  const allCategory = document.createElement("option");
  allCategory.value = "all";
  allCategory.textContent = "All policy categories";
  categorySelect.appendChild(allCategory);
  DASHBOARD_DATA.labels.forEach(label => {{
    const opt = document.createElement("option");
    opt.value = label;
    opt.textContent = label;
    categorySelect.appendChild(opt);
  }});
  [groupSelect, categorySelect, statusSelect, titleInput].forEach(el => el.addEventListener("input", render));
}}

function cleanText(text) {{ return (text || "").toString().replace(/\\s+/g, " ").trim(); }}
function isActive(row) {{ return cleanText(row["Active or Proposed Date"]).toLowerCase().startsWith("active"); }}
function isProposed(row) {{ return cleanText(row["Active or Proposed Date"]).toLowerCase().startsWith("proposed"); }}
function filteredRows() {{
  const categoryMode = categorySelect.value;
  const statusMode = statusSelect.value;
  return DASHBOARD_DATA.policies.filter(row => {{
    if (categoryMode !== "all" && row.Label !== categoryMode) return false;
    if (statusMode === "active" && !isActive(row)) return false;
    if (statusMode === "proposed" && !isProposed(row)) return false;
    return true;
  }});
}}
function visibleCountries(rows) {{
  const group = DASHBOARD_DATA.countryGroups[groupSelect.value] || [];
  const allowedCountries = new Set(group);
  const totalByCountry = new Map();
  rows.forEach(row => {{ if (!allowedCountries.has(row.Country)) return; totalByCountry.set(row.Country, (totalByCountry.get(row.Country) || 0) + 1); }});
  return group.filter(country => totalByCountry.has(country)).sort((a, b) => {{
    const diff = (totalByCountry.get(b) || 0) - (totalByCountry.get(a) || 0);
    return diff !== 0 ? diff : a.localeCompare(b);
  }});
}}
function aggregate(rows, countries) {{
  const byCountry = new Map();
  countries.forEach(country => {{ byCountry.set(country, new Map(DASHBOARD_DATA.labels.map(label => [label, []]))); }});
  rows.forEach(row => {{
    if (!byCountry.has(row.Country)) return;
    const label = row.Label;
    if (!byCountry.get(row.Country).has(label)) byCountry.get(row.Country).set(label, []);
    byCountry.get(row.Country).get(label).push(row);
  }});
  return byCountry;
}}
function niceMax(maxValue) {{ if (maxValue <= 0) return 1; const headroom = maxValue <= 1 ? 1 : 2; return Math.ceil(maxValue) + headroom; }}
function yTicks(maxValue) {{
  const top = niceMax(maxValue);
  if (top <= 10) return Array.from({{length: top + 1}}, (_, i) => i);
  const step = Math.ceil(top / 8);
  const ticks = [];
  for (let v = 0; v <= top; v += step) ticks.push(v);
  if (ticks[ticks.length - 1] !== top) ticks.push(top);
  return ticks;
}}
function labelLines(country) {{ return (DASHBOARD_DATA.displayName[country] || country).split("\\n"); }}
function truncate(text, maxLen) {{ text = cleanText(text); return text.length > maxLen ? text.slice(0, maxLen - 1) + "…" : text; }}
function escapeHTML(s) {{ return cleanText(s).replace(/[&<>"']/g, ch => ({{"&":"&amp;","<":"&lt;",">":"&gt;","\\"":"&quot;","'":"&#039;"}}[ch])); }}
function makeTooltipHTML(country, label, policies) {{
  const items = policies.map(row => `
    <li><strong>${{escapeHTML(row.Policy)}}</strong>: ${{escapeHTML(truncate(row["Policy Description"], 260))}}
      <br><span class="meta">${{escapeHTML(row["Active or Proposed Date"])}} · ${{escapeHTML(row.Source)}}</span>
    </li>`).join("");
  return `<h3>${{escapeHTML(country)}} — ${{escapeHTML(label)}}</h3><p class="meta">${{policies.length}} reform${{policies.length === 1 ? "" : "s"}} in this segment</p><ul>${{items}}</ul>`;
}}
function showTooltip(evt, html) {{ tooltip.innerHTML = html; tooltip.style.display = "block"; moveTooltip(evt); }}
function moveTooltip(evt) {{
  const pad = 16;
  const rect = tooltip.getBoundingClientRect();
  let left = evt.clientX + 14;
  let top = evt.clientY + 14;
  if (left + rect.width + pad > window.innerWidth) left = evt.clientX - rect.width - 14;
  if (top + rect.height + pad > window.innerHeight) top = evt.clientY - rect.height - 14;
  tooltip.style.left = Math.max(pad, left) + "px";
  tooltip.style.top = Math.max(pad, top) + "px";
}}
function hideTooltip() {{ tooltip.style.display = "none"; }}
function pinDetails(country, label, policies) {{
  detailCard.innerHTML = `<h2>${{escapeHTML(country)}} — ${{escapeHTML(label)}}</h2><p>Clicked segment; policy details are pinned below.</p><div class="policy-list">` +
    policies.map(row => `<div class="policy-card" style="border-left-color:${{COLORS[label] || "#999"}}"><div class="policy-title">${{escapeHTML(row.Policy)}}</div><div class="policy-desc">${{escapeHTML(row["Policy Description"])}}</div><div class="policy-foot">${{escapeHTML(row["Active or Proposed Date"])}} · ${{escapeHTML(row.Source)}}</div></div>`).join("") + "</div>";
}}
function drawLegend(activeLabels) {{
  legend.innerHTML = "";
  DASHBOARD_DATA.labels.forEach(label => {{
    if (!activeLabels.has(label)) return;
    const item = document.createElement("div");
    item.className = "legend-item";
    item.innerHTML = `<span class="legend-swatch" style="background:${{COLORS[label] || "#999"}}"></span><span>${{escapeHTML(label)}}</span>`;
    legend.appendChild(item);
  }});
}}
function render() {{
  chartTitle.textContent = titleInput.value || "";
  const rows = filteredRows();
  const countries = visibleCountries(rows);
  const byCountry = aggregate(rows, countries);
  const totals = countries.map(country => {{
    const labelMap = byCountry.get(country);
    return DASHBOARD_DATA.labels.reduce((sum, label) => sum + (labelMap.get(label) || []).length, 0);
  }});
  const maxTotal = Math.max(0, ...totals);
  const ticks = yTicks(maxTotal);
  const yMax = ticks[ticks.length - 1];
  const activeLabels = new Set();
  countries.forEach(country => {{
    const labelMap = byCountry.get(country);
    DASHBOARD_DATA.labels.forEach(label => {{ if ((labelMap.get(label) || []).length > 0) activeLabels.add(label); }});
  }});
  document.getElementById("kpiPolicies").textContent = rows.filter(r => countries.includes(r.Country)).length;
  document.getElementById("kpiCountries").textContent = countries.length;
  document.getElementById("kpiLabels").textContent = activeLabels.size;
  document.getElementById("kpiMax").textContent = maxTotal;
  subtitle.textContent = `${{groupSelect.value}} · ${{categorySelect.options[categorySelect.selectedIndex].text}} · ${{statusSelect.options[statusSelect.selectedIndex].text}}`;
  drawLegend(activeLabels);

  const countryCount = Math.max(countries.length, 1);
  const width = Math.max(920, countryCount * 82 + 130);
  const height = 560;
  const margin = {{top: 20, right: 26, bottom: 100, left: 42}};
  const plotW = width - margin.left - margin.right;
  const plotH = height - margin.top - margin.bottom;
  const barSlot = plotW / countryCount;
  const barW = Math.min(48, barSlot * 0.48);
  svg.setAttribute("viewBox", `0 0 ${{width}} ${{height}}`);
  svg.setAttribute("width", width);
  svg.setAttribute("height", height);
  svg.innerHTML = "";
  const ns = "http://www.w3.org/2000/svg";
  function el(name, attrs = {{}}, text = null) {{
    const node = document.createElementNS(ns, name);
    Object.entries(attrs).forEach(([k, v]) => node.setAttribute(k, v));
    if (text !== null) node.textContent = text;
    svg.appendChild(node);
    return node;
  }}
  function yScale(v) {{ return margin.top + plotH - (v / yMax) * plotH; }}
  ticks.forEach(t => {{
    const y = yScale(t);
    el("line", {{x1: margin.left, y1: y, x2: margin.left + plotW, y2: y, class: "axis-line", stroke: t === 0 ? "#cfcfcf" : "var(--grid)", "stroke-width": t === 0 ? 1.4 : 1.2}});
    el("text", {{x: margin.left - 18, y: y + 5, "text-anchor": "end", fill: "#666", "font-size": "15"}}, t);
  }});
  countries.forEach((country, i) => {{
    const xCenter = margin.left + i * barSlot + barSlot / 2;
    const x = xCenter - barW / 2;
    let cumulative = 0;
    const labelMap = byCountry.get(country);
    DASHBOARD_DATA.labels.forEach(label => {{
      const policies = labelMap.get(label) || [];
      const value = policies.length;
      if (value <= 0) return;
      const y1 = yScale(cumulative + value);
      const y0 = yScale(cumulative);
      const rect = el("rect", {{x: x, y: y1, width: barW, height: Math.max(1, y0 - y1), fill: COLORS[label] || "#999", class: "bar-segment", "data-country": country, "data-label": label}});
      const tip = makeTooltipHTML(country, label, policies);
      rect.addEventListener("mouseenter", e => showTooltip(e, tip));
      rect.addEventListener("mousemove", moveTooltip);
      rect.addEventListener("mouseleave", hideTooltip);
      rect.addEventListener("click", () => pinDetails(country, label, policies));
      cumulative += value;
    }});
    const lines = labelLines(country);
    const text = document.createElementNS(ns, "text");
    text.setAttribute("x", xCenter);
    text.setAttribute("y", margin.top + plotH + 24);
    text.setAttribute("text-anchor", "middle");
    text.setAttribute("fill", "#666");
    text.setAttribute("font-size", "15");
    lines.forEach((line, idx) => {{
      const tsp = document.createElementNS(ns, "tspan");
      tsp.setAttribute("x", xCenter);
      tsp.setAttribute("dy", idx === 0 ? 0 : 17);
      tsp.textContent = line;
      text.appendChild(tsp);
    }});
    svg.appendChild(text);
  }});
  el("line", {{x1: margin.left, y1: margin.top, x2: margin.left, y2: margin.top + plotH, class: "axis-line"}});
  el("line", {{x1: margin.left, y1: margin.top + plotH, x2: margin.left + plotW, y2: margin.top + plotH, class: "axis-line"}});
  if (countries.length === 0) {{ el("text", {{x: width / 2, y: height / 2, "text-anchor": "middle", fill: "#777", "font-size": "18"}}, "No rows match the current filters."); }}
}}
initControls();
render();
</script>
</body>
</html>"""


def find_workbook(input_dir: Path, region_key: str) -> Path:
    """Locate ``<region>.xlsx`` (preferred) or any ``*<region>*.xlsx`` fallback."""
    canonical = input_dir / f"{region_key}.xlsx"
    if canonical.exists():
        return canonical
    matches = sorted(
        (
            p
            for p in input_dir.glob(f"*{region_key}*.xlsx")
            if not p.name.startswith("~$")
        ),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if not matches:
        raise FileNotFoundError(
            f"No workbook for region '{region_key}' in {input_dir}. "
            f"Expected {canonical.name} or any *{region_key}*.xlsx."
        )
    if len(matches) > 1:
        print(
            f"  multiple files matched '*{region_key}*.xlsx'; using most recent: {matches[0].name}"
        )
    return matches[0]


def generate_region(
    region_cfg: Dict[str, Any],
    input_dir: Path,
    output_dir: Path,
    chart_title: str,
    tracker: Optional[str] = None,
) -> Dict[str, Any]:
    key = region_cfg["key"]
    display = region_cfg.get("display_name", key)
    print(f"[{key}] {display}")
    workbook = find_workbook(input_dir, key)
    print(f"  workbook: {workbook.name}")
    rows, sheet_name, header_row = load_policy_rows(
        workbook, region_cfg.get("sheet", CANONICAL_SHEET)
    )
    print(
        f"  sheet: {sheet_name}; header row: {header_row}; raw policy-like rows: {len(rows)}"
    )
    rows, excluded = apply_region_filters(rows, region_cfg)
    print(
        f"  included rows: {len(rows)}; excluded rows: {excluded}; countries/economies: {len({r['Country'] for r in rows})}"
    )

    from text.plotting.policy_dashboards_v6 import (
        build_v6_dashboard_data,
        has_v6_columns,
        make_v6_html,
    )

    out_dir = tracker_dir(output_dir, tracker)
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / addon_filename(key)

    if has_v6_columns(rows):
        print("  taxonomy: v6 (Category + Subcategory)")
        groups = build_country_groups(rows, region_cfg)
        display_names = build_display_names(sort_countries_by_count(rows))
        data, colors = build_v6_dashboard_data(
            rows, region_cfg, workbook, sheet_name, excluded, groups, display_names
        )
        html = make_v6_html(
            data,
            chart_title=chart_title,
            page_title=f"{tracker_label(tracker)} Dashboard",
            chart_aria_subject=get_tracker(tracker)["aria_subject"],
        )
        labels_summary = data["categories"]
    else:
        print("  taxonomy: legacy Label")
        data, colors = build_dashboard_data(
            rows, region_cfg, workbook, sheet_name, excluded
        )
        html = make_html(data, colors, chart_title=chart_title)
        labels_summary = data["labels"]

    out_path.write_text(html, encoding="utf-8")
    print(f"  wrote: {out_path}")
    return {
        "key": key,
        "display_name": display,
        "workbook": str(workbook),
        "output": str(out_path),
        "included_rows": len(rows),
        "excluded_rows": excluded,
        "countries": len({r["Country"] for r in rows}),
        "labels": labels_summary,
    }


def _resolve_regions(only: Optional[Iterable[str]]) -> List[Dict[str, Any]]:
    if not only:
        return REGIONS
    requested = {x.lower() for x in only}
    selected = [r for r in REGIONS if r["key"] in requested]
    missing = requested - {r["key"] for r in selected}
    if missing:
        raise ValueError(f"Unknown region key(s): {', '.join(sorted(missing))}")
    return selected


def build_addons(
    region: Optional[str] = None,
    chart_title: Optional[str] = None,
    input_dir: Optional[Path] = None,
    output_dir: Optional[Path] = None,
    tracker: Optional[str] = None,
) -> List[Dict[str, Any]]:
    """Programmatic entry point used by the ``po text build-policy-addons`` CLI."""
    in_dir = workbook_dir(input_dir or DEFAULT_INPUT_DIR, tracker)
    out_dir = output_dir or DEFAULT_OUTPUT_DIR
    if not in_dir.exists():
        raise FileNotFoundError(f"Input directory does not exist: {in_dir}")
    title = chart_title or dt.date.today().strftime("%-d %B %Y")
    only = [region] if region else None
    regions = _resolve_regions(only)
    summary: List[Dict[str, Any]] = []
    for region_cfg in regions:
        summary.append(generate_region(region_cfg, in_dir, out_dir, title, tracker))
    return summary


def make_zip(output_dir: Path, zip_path: Path) -> None:
    html_files = sorted(output_dir.glob("*.html"))
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for path in html_files:
            zf.write(path, arcname=path.name)
    print(f"Created zip: {zip_path}")


def parse_args(argv: Optional[List[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build regional Fuel Crisis Policy addon HTMLs from policy_tracker workbooks."
    )
    parser.add_argument(
        "--input-dir",
        type=Path,
        default=DEFAULT_INPUT_DIR,
        help=f"Folder with <region>.xlsx workbooks. Default: {DEFAULT_INPUT_DIR}",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=DEFAULT_OUTPUT_DIR,
        help=f"Folder for addon HTMLs. Default: {DEFAULT_OUTPUT_DIR}",
    )
    parser.add_argument(
        "--only",
        nargs="*",
        help="Region keys to build, e.g. --only ssa sar. Omit to build all.",
    )
    parser.add_argument(
        "--chart-title",
        default=None,
        help="Editable chart title embedded in each HTML. Defaults to today's date.",
    )
    parser.add_argument(
        "--tracker",
        default=DEFAULT_TRACKER,
        choices=sorted(TRACKERS),
        help="Policy-tracker variant to build. Default: fuel.",
    )
    parser.add_argument(
        "--zip",
        action="store_true",
        help="Also zip all generated HTML files into output-dir.",
    )
    parser.add_argument(
        "--zip-file",
        type=Path,
        default=None,
        help="Optional zip filename. Defaults to <output-dir>/regional_fuel_crisis_policy_dashboards_html.zip",
    )
    return parser.parse_args(argv)


def main(argv: Optional[List[str]] = None) -> int:
    args = parse_args(argv)
    in_dir = workbook_dir(args.input_dir, args.tracker)
    if not in_dir.exists():
        print(f"Input directory does not exist: {in_dir}", file=sys.stderr)
        return 2
    title = args.chart_title or dt.date.today().strftime("%-d %B %Y")

    try:
        regions = _resolve_regions(args.only)
    except ValueError as exc:
        print(str(exc), file=sys.stderr)
        return 2

    summary = []
    errors = []
    for region_cfg in regions:
        try:
            summary.append(
                generate_region(
                    region_cfg, in_dir, args.output_dir, title, args.tracker
                )
            )
        except Exception as exc:
            errors.append((region_cfg["key"], str(exc)))
            print(f"  ERROR: {exc}", file=sys.stderr)

    args.output_dir.mkdir(parents=True, exist_ok=True)
    summary_path = args.output_dir / "dashboard_generation_summary.json"
    summary_path.write_text(
        json.dumps(
            {"generated": summary, "errors": errors}, indent=2, ensure_ascii=False
        ),
        encoding="utf-8",
    )
    print(f"Wrote summary: {summary_path}")

    if args.zip:
        zip_path = args.zip_file or (
            args.output_dir / "regional_fuel_crisis_policy_dashboards_html.zip"
        )
        make_zip(args.output_dir, zip_path)

    if errors:
        print("\nSome regions failed:", file=sys.stderr)
        for key, err in errors:
            print(f"- {key}: {err}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
