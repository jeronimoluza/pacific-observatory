"""Malawi National Statistics Office (NSO) -- monthly Consumer Price Index.

NSO publishes a monthly "CPI Stats Flash" news post at nsomalawi.mw (a Nuxt
SSR site backed by a headless Laravel CMS at cms.nsomalawi.mw). Rather than
hardcode a slug/date, this fetcher discovers the latest release live:

1. GET /news -- the listing page's Nuxt payload embeds slugs for the last
   several months' posts as plain text, e.g.
   "consumer-price-index-july-2026". Regex-extract all such slugs and pick
   the most recent by (year, month).
2. GET /news/<slug> -- the post's own Nuxt payload lists its attachments
   (xlsx/pdf) just before a "popular:[" related-posts block that embeds
   *other* posts' attachments too; we truncate the page at "popular:[" so
   we only match the target post's own files, not a neighbour's.
3. Pick the attachment named "NSO-CPI-<Month>-<Year>---Stats-Flash.xlsx"
   (there is a sibling "YoY-NSO-Stats-Flash" file we skip -- year-on-year
   inflation rates, not the index-level table).
4. Parse "Sheet1" of that xlsx directly with openpyxl. Layout (verified
   2026-08-31 against the July 2026 release): a title row, a header row
   (columns F..R = Food, Alcohol & Tobacco, Clothing & footwear,
   Housing/Water/Electricity, Furnishing & Household, Health, Transport,
   Communication, Recreation & Culture, Education, Restaurants & Hotels,
   Miscellaneous, All items), then three stacked blocks -- "National",
   "Urban", "Rural" -- each starting with a weights row (label + weights,
   ignored here), a "Dec, 2021" base-100 row, and then one row per month
   going backwards in time (year only stamped on the first row of each
   year; carried forward for the rest). Base period is "Dec 2021=100"
   (read from the title row; not hardcoded beyond the fallback).

COICOP-2018 mapping mirrors the convention already used by
eap/pacific_islands/vanuatu/vnso_cpi.py: the 11 divisions NSO breaks out
map 1:1 to COICOP 01-11, and NSO's "Miscellaneous" (which absorbs personal
care / other services not broken out separately, same as Vanuatu's own
"Miscellaneous" column) maps to COICOP 13 ("Personal care, social
protection and miscellaneous goods and services") rather than 12
("Insurance and financial services"). The headline "All items" column is
dropped -- no sanctioned headline sentinel yet, per the same open design
question noted in vnso_cpi.py. "Month to month/Food/Non-Food Inflation"
percentage columns (C/D/E) are also dropped -- they are rates, not index
levels, and are trivially re-derivable from the index series if needed.

All three location blocks are emitted; `subnational_area` is None for the
national series and "Urban"/"Rural" for the other two, so same-month/
same-division rows across locations don't collide on the identifying
tuple.
"""

from __future__ import annotations

import io
import logging
import re
from datetime import date

import openpyxl
import pandas as pd

from prices.fetchers.utils import get_scrape_ts, get_session, make_hash

logger = logging.getLogger(__name__)

_NEWS_LIST_URL = "https://nsomalawi.mw/news"
_POST_URL_TMPL = "https://nsomalawi.mw/news/{slug}"
_COUNTRY = "Malawi"
_SOURCE_KEY = "mwi_nso_cpi"
_DEFAULT_BASE_PERIOD = "Dec2021=100"
_IDENT = ["source_key", "observation_date", "coicop_code", "subnational_area"]

_SLUG_RE = re.compile(r"consumer-price-index-([a-z]+)-(\d{4})")
_ATTACHMENT_RE = re.compile(r'url:"([^"]*?Stats-Flash[^"]*?\.xlsx)"')

_MONTH_NUM = {
    "january": 1,
    "jan": 1,
    "february": 2,
    "feb": 2,
    "march": 3,
    "mar": 3,
    "april": 4,
    "apr": 4,
    "may": 5,
    "june": 6,
    "jun": 6,
    "july": 7,
    "jul": 7,
    "august": 8,
    "aug": 8,
    "september": 9,
    "sep": 9,
    "sept": 9,
    "october": 10,
    "oct": 10,
    "november": 11,
    "nov": 11,
    "december": 12,
    "dec": 12,
}

# Column letter -> COICOP-2018 division, per the docstring's mapping.
_DIVISION_COLS = {
    6: "01",  # F: Food
    7: "02",  # G: Alcoholic drinks & Tobacco
    8: "03",  # H: Clothing & footwear
    9: "04",  # I: Housing, Water & Electricity
    10: "05",  # J: Furnishing & Household
    11: "06",  # K: Health
    12: "07",  # L: Transportation
    13: "08",  # M: Communication
    14: "09",  # N: Recreation & Culture
    15: "10",  # O: Education
    16: "11",  # P: Restaurants and Hotels
    17: "13",  # Q: Miscellaneous -> COICOP 13, mirrors vnso_cpi.py
    # R (18, "All items") deliberately dropped -- headline, no sentinel.
}

_BLOCK_LABELS = {"national", "urban", "rural"}


def _find_latest_slug(html: str) -> tuple[str, int, int] | None:
    """Returns (slug, year, month) for the most recent CPI post found on the
    news listing page, or None if no slug matched."""
    best: tuple[int, int, str] | None = None
    for month_name, year_str in set(_SLUG_RE.findall(html)):
        month = _MONTH_NUM.get(month_name.lower())
        if month is None:
            continue
        year = int(year_str)
        if best is None or (year, month) > (best[0], best[1]):
            best = (year, month, f"consumer-price-index-{month_name}-{year_str}")
    if best is None:
        return None
    year, month, slug = best
    return slug, year, month


def _find_stats_flash_xlsx(html: str) -> str | None:
    """Returns the CPI-levels xlsx URL for THIS post (not a related post
    embedded further down the page), or None if not found."""
    boundary = html.find("popular:[")
    scope = html if boundary == -1 else html[:boundary]
    candidates = [m.replace("\\u002F", "/") for m in _ATTACHMENT_RE.findall(scope)]
    # Skip the year-on-year variant; keep the index-levels file.
    levels = [c for c in candidates if "yoy" not in c.lower()]
    return levels[-1] if levels else None


def _parse_sheet(ws) -> list[tuple[str | None, date, str, float]]:
    """Returns (subnational_area, obs_date, coicop_code, index_value) rows."""
    results: list[tuple[str | None, date, str, float]] = []
    current_area: str | None = None
    current_year: int | None = None
    seen_any_block = False

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        a = row[0].value
        b = row[1].value if len(row) > 1 else None

        if isinstance(a, str) and a.strip().lower() in _BLOCK_LABELS:
            current_area = None if a.strip().lower() == "national" else a.strip()
            current_year = None
            seen_any_block = True
            continue
        if not seen_any_block:
            continue
        if isinstance(a, str) and a.strip().lower() == "weight":
            continue

        month_num: int | None = None
        if isinstance(a, str) and a.strip().lower().startswith("dec,"):
            current_year = 2021
            month_num = 12
        elif isinstance(a, (int, float)):
            current_year = int(a)
            if isinstance(b, str):
                month_num = _MONTH_NUM.get(b.strip().lower())
        elif a is None and isinstance(b, str):
            month_num = _MONTH_NUM.get(b.strip().lower())

        if month_num is None or current_year is None:
            continue

        obs_date = date(current_year, month_num, 1)
        for col_idx, coicop in _DIVISION_COLS.items():
            cell = row[col_idx - 1] if col_idx - 1 < len(row) else None
            val = cell.value if cell is not None else None
            if val is None:
                continue
            try:
                idx_val = float(val)
            except (TypeError, ValueError):
                continue
            results.append((current_area, obs_date, coicop, idx_val))

    return results


def fetch_mwi_nso_cpi(cutoff: date) -> pd.DataFrame | None:
    session = get_session()

    resp = session.get(_NEWS_LIST_URL, timeout=30)
    resp.raise_for_status()
    found = _find_latest_slug(resp.text)
    if found is None:
        logger.warning(
            "[%s] Could not find any consumer-price-index-* slug on %s",
            _SOURCE_KEY,
            _NEWS_LIST_URL,
        )
        return None
    slug, year, month = found

    post_url = _POST_URL_TMPL.format(slug=slug)
    post_resp = session.get(post_url, timeout=30)
    post_resp.raise_for_status()
    xlsx_url = _find_stats_flash_xlsx(post_resp.text)
    if xlsx_url is None:
        logger.warning(
            "[%s] Could not find a Stats-Flash xlsx attachment on %s",
            _SOURCE_KEY,
            post_url,
        )
        return None

    xlsx_resp = session.get(xlsx_url, timeout=60)
    xlsx_resp.raise_for_status()
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_resp.content), data_only=True)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.worksheets[0]

    parsed = _parse_sheet(ws)
    if not parsed:
        logger.warning("[%s] Parsed zero rows from %s", _SOURCE_KEY, xlsx_url)
        return None

    rows = []
    for subnational_area, obs_date, coicop, idx_val in parsed:
        if obs_date <= cutoff:
            continue
        row = {
            "observation_date": obs_date.isoformat(),
            "period_kind": "monthly_avg",
            "country": _COUNTRY,
            "subnational_area": subnational_area,
            "source_key": _SOURCE_KEY,
            "coicop_code": coicop,
            "index_value": idx_val,
            "index_base_period": _DEFAULT_BASE_PERIOD,
            "source_url": xlsx_url,
            "notes": f"NSO CPI Stats Flash, post={post_url}",
            "scrape_ts": get_scrape_ts(),
            "observation_hash": None,
        }
        row["observation_hash"] = make_hash(row, _IDENT)
        rows.append(row)

    return pd.DataFrame(rows) if rows else None
