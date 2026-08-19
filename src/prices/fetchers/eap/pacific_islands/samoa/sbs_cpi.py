"""Samoa Bureau of Statistics (sbs.gov.ws) — Consumer Price Index, monthly.

SBS publishes a monthly CPI release (sbs.gov.ws/cpi/) as a paired PDF + XLSX
under /documents/economics/CPI/<year>/ (older releases use /images/sbs-
documents/Economics/CPI/<year>/, not walked here). Verified live 2026-08-11:
the latest XLSX has 7 sheets; only "Table 1&2" is used here — "Table 1.
CONSUMER PRICE INDEX" (base period: average prices Feb 2016 = 100). Its layout
is ROWS = time period, COLUMNS = named COICOP-ish division (a much simpler
shape than the same workbook's "Table 3. AVERAGE RETAIL PRICES OF SELECTED
COMMODITIES" sheet, which is a genuinely rich ~190-item whole-basket
official_avg series across all 13 divisions with real commodity names
("Chicken leg quarters", "Bread", "Milk Anchor", ...) — NOT built this round
because its header uses a 2-row merged year/month layout that needs careful
merged-cell resolution; flagged in the onboarding report as a strong follow-up
target, arguably richer than the Local Market Survey).

Row structure for Table 1&2: a bare 4-digit year in column B ("2025", "2026")
marks a new year section; subsequent rows are month names ("May ", "June  (P)"
— "(P)" = provisional, stripped) until the next year marker or a non-date row
("Weights", "Ave - 2020"..."Ave - 2025" annual summaries, "Percentage Change",
"Source :" footer). Only genuine monthly rows are emitted — the annual "Ave -
YYYY" summary rows are dropped (same resolution as every other monthly
cpi_benchmark fetcher in this codebase; they're a derived re-aggregation, not
a distinct observation). Column headers (row 4) name the division directly
("Food and Non Alcoholic Beverages", "Health", ...) rather than a numeric
code — mapped to COICOP-2018 divisions below. "All Items" (headline) is
dropped, no sanctioned sentinel (open design question in the onboarding
skill). Division 13 (insurance/financial services) is absent from Samoa's own
basket, same 12-division pattern seen in Solomon Islands' SINSO CPI and
Indonesia's BPS CPI in this codebase.

Emits IndexObservation rows (analytical_role: cpi_benchmark).
coicop_classification: publisher_labeled (static _COICOP_MAP below).
"""

from __future__ import annotations

import io
import logging
import re
from datetime import date

import openpyxl
import pandas as pd

from prices.fetchers.utils import get_scrape_ts, get_session, make_hash

logger = logging.getLogger(__name__)

_COUNTRY = "Samoa"
_SOURCE_KEY = "ws_sbs_cpi"
_IDENT = ["source_key", "observation_date", "coicop_code"]
_BASE_PERIOD = "Feb2016=100"
_CPI_URL = "https://www.sbs.gov.ws/cpi/"
_XLSX_HREF_RE = re.compile(
    r'href="(https://www\.sbs\.gov\.ws/documents/economics/CPI/(\d{4})/[^"]*\.xlsx)"',
    re.IGNORECASE,
)

_COICOP_MAP = {
    "Food and Non Alcoholic Beverages": "01",
    "Alcoholic Beverages, Tobacco and Narcotics": "02",
    "Clothing and Footwear": "03",
    "Housing,Water, Electricity, Gas and other Fuels": "04",
    "Furnishings, Household Equipment, and Maintenance": "05",
    "Health": "06",
    "Transport": "07",
    "Communication": "08",
    "Recreation and Culture": "09",
    "Education": "10",
    "Restaurants": "11",
    "Miscellaneous goods and services": "12",
}

_MONTHS = {
    "JAN": 1,
    "FEB": 2,
    "MAR": 3,
    "APR": 4,
    "MAY": 5,
    "JUN": 6,
    "JUL": 7,
    "AUG": 8,
    "SEP": 9,
    "OCT": 10,
    "NOV": 11,
    "DEC": 12,
}
_STOP_MARKERS = ("Percentage Change", "Source", "Weights", "Ave -", "Ave-")


def _find_latest_xlsx_url(session) -> str | None:
    try:
        resp = session.get(_CPI_URL, timeout=30)
        resp.raise_for_status()
    except Exception as exc:  # noqa: BLE001
        logger.warning("[%s] CPI page fetch failed: %s", _SOURCE_KEY, exc)
        return None
    matches = _XLSX_HREF_RE.findall(resp.text)
    if not matches:
        return None

    def _month_key(item: tuple[str, str]) -> tuple[int, int]:
        url, year = item
        base = url.rsplit("/", 1)[-1]
        m = re.match(r"0?(\d{1,2})[-_]", base)
        month = int(m.group(1)) if m else 0
        return (int(year), month)

    return max(matches, key=_month_key)[0]


def _row_date(label: object, current_year: int | None) -> date | None:
    if current_year is None or not isinstance(label, str):
        return None
    cleaned = re.sub(r"\(P\)", "", label, flags=re.IGNORECASE).strip()
    if len(cleaned) < 3:
        return None
    key = cleaned[:3].upper()
    month = _MONTHS.get(key)
    return date(current_year, month, 1) if month else None


def _rows_from_xlsx(xlsx_bytes: bytes, url: str, cutoff: date) -> list[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb["Table 1&2"]
    header = [
        h.strip() if isinstance(h, str) else h
        for h in (ws.cell(row=4, column=c).value for c in range(1, ws.max_column + 1))
    ]
    col_codes = [
        (c, _COICOP_MAP[header[c - 1]])
        for c in range(1, len(header) + 1)
        if header[c - 1] in _COICOP_MAP
    ]
    if not col_codes:
        return []

    ts_scrape = get_scrape_ts()
    out: list[dict] = []
    current_year: int | None = None
    for r in range(5, ws.max_row + 1):
        label = ws.cell(row=r, column=2).value
        if isinstance(label, str) and any(m in label for m in _STOP_MARKERS):
            continue
        if isinstance(label, (int, float)) and 2000 <= int(label) <= 2100:
            current_year = int(label)
            continue
        if (
            isinstance(label, str)
            and label.strip().isdigit()
            and len(label.strip()) == 4
        ):
            current_year = int(label.strip())
            continue
        obs_date = _row_date(label, current_year)
        if obs_date is None:
            continue
        if obs_date <= cutoff:
            continue
        for col, coicop in col_codes:
            val = ws.cell(row=r, column=col).value
            try:
                index_value = float(val)
            except (TypeError, ValueError):
                continue
            rec = {
                "observation_date": obs_date.isoformat(),
                "period_kind": "monthly_avg",
                "country": _COUNTRY,
                "source_key": _SOURCE_KEY,
                "coicop_code": coicop,
                "index_value": round(index_value, 4),
                "index_base_period": _BASE_PERIOD,
                "source_url": url,
                "notes": f"category={header[col - 1]}",
                "scrape_ts": ts_scrape,
                "observation_hash": None,
            }
            rec["observation_hash"] = make_hash(rec, _IDENT)
            out.append(rec)
    return out


def fetch_ws_sbs_cpi(cutoff: date) -> pd.DataFrame | None:
    session = get_session()
    session.headers.update(
        {
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) Chrome/120 Safari/537.36"
        }
    )
    xlsx_url = _find_latest_xlsx_url(session)
    if not xlsx_url:
        logger.warning("[%s] no CPI xlsx link found", _SOURCE_KEY)
        return None
    try:
        resp = session.get(xlsx_url, timeout=60)
        resp.raise_for_status()
    except Exception as exc:  # noqa: BLE001
        logger.warning("[%s] xlsx fetch failed: %s", _SOURCE_KEY, exc)
        return None
    rows = _rows_from_xlsx(resp.content, xlsx_url, cutoff)
    logger.info("[%s] %d rows (cutoff=%s)", _SOURCE_KEY, len(rows), cutoff)
    return pd.DataFrame(rows) if rows else None
