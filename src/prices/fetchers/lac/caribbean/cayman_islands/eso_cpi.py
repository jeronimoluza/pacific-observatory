"""Cayman Islands Economics and Statistics Office (eso.ky) — Consumer Price
Index, quarterly, by major COICOP-ish group.

ESO publishes its indicator downloads on a single static landing page
(eso.ky/indicators-page.html) whose "CPI by Divisons <Nq><YYYY>.xlsx" link is
overwritten in place each quarter with a filename that embeds the latest
quarter tag (verified live 2026-09-01: "CPI by Divisons 4q2025.xlsx"). The
fetcher re-resolves the link from the landing page each run rather than
hardcoding a quarter tag or a numeric storage id (both change every release).

The workbook (single sheet "Table 4") is NOT one continuous table — it
stacks TWO title blocks top to bottom, because ESO rebased the index
partway through the series: rows 1-83 are "(JUNE 2008 = 100)" covering
2009–2017, and a second "Table 4" title at row 84 restarts the same
2009-onward history "(SEPTEMBER 2016 = 100)", extending through the current
quarter. Both blocks cover overlapping years at DIFFERENT index levels (the
same physical quarter has two different numbers depending which base it's
read against) — treating the sheet as one continuous stream double-counts
every overlapping quarter with two conflicting values sharing one
(date, coicop_code) identity. Anchor on the LAST "Table 4" title block only
(same "anchor on the LAST occurrence" rule used for regulator PDFs with a
superseded earlier table) — verified live: only two such blocks exist, and
the second is both the most recent base period AND the one with full
coverage through the present quarter, so no data is lost by dropping the
first.

Within the (single, latest) block: a header row names 12 major groups +
"ALL ITEMS", a WEIGHT row, and then repeating year blocks: a bare 4-digit
year in column C marks a new year, followed by
"MARCH"/"JUNE"/"SEPTEMBER"/"DECEMBER" quarterly rows, an "ANNUAL AVERAGE
<year>" row (dropped — a derived re-aggregation of the four quarters, same
convention as every other cpi_benchmark fetcher in this codebase), and a
blank separator. The tail of the block repeats the same shape for "ANNUAL
AVERAGE % CHANGE <year>" plus two trailing "% CHANGE OVER PREV YEAR"/"%
CHANGE OVER PREV QTR" rows — all dropped, they're derived statistics, not
index levels. Coverage verified live within the anchored block: 2009 Q1
through 2025 Q4 (68 quarters), quarterly cadence throughout — Cayman's CPI
is a genuinely quarterly series (not monthly).

"ALL ITEMS" (headline) is dropped — no sanctioned coicop_code sentinel yet
(open design question in the onboarding skill). Division 13 (insurance and
financial services) is absent from Cayman's own basket — this is a 12-group
COICOP-1999-style classification, same pattern already seen on Samoa (SBS),
Solomon Islands (SINSO) and Indonesia (BPS) in this codebase.

Emits IndexObservation rows (analytical_role: cpi_benchmark).
coicop_classification: publisher_labeled (static _COICOP_MAP below).
"""

from __future__ import annotations

import io
import logging
import re
from datetime import date

import openpyxl
import pandas as pd

from prices.fetchers.utils import get_scrape_ts, get_session, make_hash

logger = logging.getLogger(__name__)

_COUNTRY = "Cayman Islands"
_SOURCE_KEY = "ky_eso_cpi"
_IDENT = ["source_key", "observation_date", "coicop_code"]
_BASE_PERIOD = "Sep2016=100"  # fallback only; the real value is read from the workbook's latest title block
_INDICATORS_URL = "https://www.eso.ky/indicators-page.html"
_XLSX_HREF_RE = re.compile(
    r'href="(/storage/indicator_docums/uploadFilePdf/\d+/CPI\s+by\s+Divisons?[^"]*\.xlsx)"',
    re.IGNORECASE,
)

_COICOP_MAP = {
    "Food & Non-alcoholic beverages": "01",
    "Alcoholic Beverages & Tobacco": "02",
    "Clothing & Footwear": "03",
    "Housing and Utilities": "04",
    "Household Equipment": "05",
    # Renamed in the Sept-2016-rebased block (row 87) from "Household
    # Equipment" to "Household Furnishings & Equipment" — same division,
    # different label; verified live 2026-09-01.
    "Household Furnishings & Equipment": "05",
    "Health": "06",
    "Transport": "07",
    "Communication": "08",
    "Recreation & Culture": "09",
    "Education": "10",
    "Restaurants & Hotels": "11",
    "Miscellaneous Goods & Services": "12",
}

_MONTH_TO_QTR = {"MARCH": 3, "JUNE": 6, "SEPTEMBER": 9, "DECEMBER": 12}
_STOP_MARKERS = (
    "ANNUAL AVERAGE",
    "% CHANGE OVER PREV",
    "WEIGHT",
)


def _find_xlsx_url(session) -> str | None:
    try:
        resp = session.get(_INDICATORS_URL, timeout=30)
        resp.raise_for_status()
    except Exception as exc:  # noqa: BLE001
        logger.warning("[%s] indicators page fetch failed: %s", _SOURCE_KEY, exc)
        return None
    m = _XLSX_HREF_RE.search(resp.text)
    if not m:
        return None
    return "https://www.eso.ky" + m.group(1).replace(" ", "%20")


_BASE_PERIOD_RE = re.compile(r"\(([^()]*=\s*100)\)")


def _rows_from_xlsx(xlsx_bytes: bytes, url: str, cutoff: date) -> list[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]

    # The sheet stacks one or more "PERIOD / DIVISION" header blocks (one
    # per historical rebase — see module docstring). Anchor on the LAST
    # one: it's both the current base period and the block with full
    # coverage through the present quarter.
    header_row = None
    label_col = None
    base_period = _BASE_PERIOD
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and "PERIOD" in v.upper() and "DIVISION" in v.upper():
                header_row = r
                label_col = c
                # base-period text lives a row or two above the header,
                # inside the "(... = 100)" title line.
                for back in range(1, 4):
                    tv = ws.cell(row=r - back, column=c).value
                    if isinstance(tv, str):
                        bm = _BASE_PERIOD_RE.search(tv)
                        if bm:
                            base_period = bm.group(1).replace(" ", "")
                            break
                break
    if header_row is None:
        logger.warning("[%s] header row not found", _SOURCE_KEY)
        return []

    header = [
        h.strip() if isinstance(h, str) else h
        for h in (
            ws.cell(row=header_row, column=c).value for c in range(1, ws.max_column + 1)
        )
    ]
    col_codes = [
        (c, _COICOP_MAP[header[c - 1]])
        for c in range(1, len(header) + 1)
        if isinstance(header[c - 1], str) and header[c - 1] in _COICOP_MAP
    ]
    if not col_codes:
        return []

    ts_scrape = get_scrape_ts()
    out: list[dict] = []
    current_year: int | None = None
    for r in range(header_row + 1, ws.max_row + 1):
        label = ws.cell(row=r, column=label_col).value
        if isinstance(label, str) and any(m in label.upper() for m in _STOP_MARKERS):
            continue
        if isinstance(label, (int, float)) and 2000 <= int(label) <= 2100:
            current_year = int(label)
            continue
        if (
            isinstance(label, str)
            and label.strip().isdigit()
            and len(label.strip()) == 4
        ):
            current_year = int(label.strip())
            continue
        if not isinstance(label, str) or current_year is None:
            continue
        month = _MONTH_TO_QTR.get(label.strip().upper())
        if month is None:
            continue
        obs_date = date(current_year, month, 1)
        if obs_date <= cutoff:
            continue
        for col, coicop in col_codes:
            val = ws.cell(row=r, column=col).value
            try:
                index_value = float(val)
            except (TypeError, ValueError):
                continue
            rec = {
                "observation_date": obs_date.isoformat(),
                "period_kind": "quarterly_avg",
                "country": _COUNTRY,
                "source_key": _SOURCE_KEY,
                "coicop_code": coicop,
                "index_value": round(index_value, 4),
                "index_base_period": base_period,
                "source_url": url,
                "notes": f"category={header[col - 1]}",
                "scrape_ts": ts_scrape,
                "observation_hash": None,
            }
            rec["observation_hash"] = make_hash(rec, _IDENT)
            out.append(rec)
    return out


def fetch_ky_eso_cpi(cutoff: date) -> pd.DataFrame | None:
    session = get_session()
    session.headers.update(
        {
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        }
    )
    xlsx_url = _find_xlsx_url(session)
    if not xlsx_url:
        logger.warning("[%s] no CPI-by-divisions xlsx link found", _SOURCE_KEY)
        return None
    try:
        resp = session.get(xlsx_url, timeout=60)
        resp.raise_for_status()
    except Exception as exc:  # noqa: BLE001
        logger.warning("[%s] xlsx fetch failed: %s", _SOURCE_KEY, exc)
        return None
    rows = _rows_from_xlsx(resp.content, xlsx_url, cutoff)
    logger.info("[%s] %d rows (cutoff=%s)", _SOURCE_KEY, len(rows), cutoff)
    return pd.DataFrame(rows) if rows else None
