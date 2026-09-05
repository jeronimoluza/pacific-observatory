"""Build an Excel catalog of fuel-price sources present in retail_series_enriched.csv.

Reads metadata from FETCHER_REGISTRY + SOURCE_META (per-fetcher modules) and
observation date ranges from the enriched CSV.  Writes scripts/source_catalog.xlsx.

Usage:
    poetry run python scripts/sources_spreadsheet.py
"""

from __future__ import annotations

import csv
import importlib
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

from src.cpi.fuel_prices.fetchers import FETCHER_REGISTRY  # noqa: E402

ENRICHED_CSV = (
    PROJECT_ROOT
    / "data"
    / "cpi"
    / "fuel_prices_staged"
    / "enrich"
    / "retail_series_enriched.csv"
)
OUTPUT_XLSX = PROJECT_ROOT / "scripts" / "source_catalog.xlsx"

# Fetcher modules to scan for SOURCE_META
_FETCHER_MODULES = [
    "src.cpi.fuel_prices.fetchers.australia",
    "src.cpi.fuel_prices.fetchers.cambodia",
    "src.cpi.fuel_prices.fetchers.china",
    "src.cpi.fuel_prices.fetchers.fiji",
    "src.cpi.fuel_prices.fetchers.global_commodities",
    "src.cpi.fuel_prices.fetchers.hong_kong",
    "src.cpi.fuel_prices.fetchers.indonesia",
    "src.cpi.fuel_prices.fetchers.japan",
    "src.cpi.fuel_prices.fetchers.korea",
    "src.cpi.fuel_prices.fetchers.lao",
    "src.cpi.fuel_prices.fetchers.malaysia",
    "src.cpi.fuel_prices.fetchers.mongolia",
    "src.cpi.fuel_prices.fetchers.myanmar",
    "src.cpi.fuel_prices.fetchers.new_zealand",
    "src.cpi.fuel_prices.fetchers.pacific_islands",
    "src.cpi.fuel_prices.fetchers.philippines",
    "src.cpi.fuel_prices.fetchers.singapore",
    "src.cpi.fuel_prices.fetchers.taiwan",
    "src.cpi.fuel_prices.fetchers.thailand",
    "src.cpi.fuel_prices.fetchers.timor_leste",
    "src.cpi.fuel_prices.fetchers.tonga",
    "src.cpi.fuel_prices.fetchers.vietnam",
]

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _simplify_extraction_method(methods: list[str] | str) -> str:
    """Map extraction_method to a human-friendly data-format label.

    Classifies each method individually to avoid substring false-positives
    (e.g. "scraping" contains "api").
    """
    if isinstance(methods, str):
        methods = [methods]
    if not methods:
        return "Unknown"

    tags: set[str] = set()
    for m in methods:
        ml = m.lower().strip()
        if "excel" in ml:
            tags.add("excel")
        if "pdf" in ml:
            tags.add("pdf")
        if "csv" in ml or "parquet" in ml:
            tags.add("csv")
        if "rss" in ml or ml == "xml":
            tags.add("rss")
        if "soap" in ml:
            tags.add("soap")
        if "rest" in ml or "json" in ml or "firebase" in ml:
            tags.add("json_api")
        if "ckan" in ml:
            tags.add("csv")
        if ml == "api":
            tags.add("json_api")
        if "ocr" in ml or "image" in ml:
            tags.add("ocr")
        if "html table" in ml:
            tags.add("html_table")

    if "excel" in tags:
        return "Excel"
    if "pdf" in tags and "ocr" in tags:
        return "PDF (OCR)"
    if "pdf" in tags:
        return "PDF"
    if "csv" in tags:
        return "CSV"
    if "rss" in tags:
        return "RSS/XML"
    if "soap" in tags:
        return "SOAP API"
    if "json_api" in tags:
        return "JSON API"
    if "html_table" in tags:
        return "HTML table"
    if "ocr" in tags:
        return "Web scraping (OCR)"
    return "Web scraping"


def _collect_source_meta() -> dict[str, dict[str, object]]:
    """Import fetcher modules → {source_key: {url, extraction_method}}."""
    meta_by_key: dict[str, dict[str, object]] = {}
    for mod_name in _FETCHER_MODULES:
        try:
            mod = importlib.import_module(mod_name)
        except Exception as exc:
            print(f"  WARNING: could not import {mod_name}: {exc}")
            continue
        meta_list = getattr(mod, "SOURCE_META", None)
        if not meta_list:
            continue
        for entry in meta_list:
            source_keys = entry.get("source_keys", [])
            if isinstance(source_keys, str):
                source_keys = [source_keys]
            if not source_keys:
                continue
            url = entry.get("url", "")
            extraction_method = entry.get("extraction_method", [])
            for sk in source_keys:
                if sk not in meta_by_key:
                    meta_by_key[sk] = {
                        "url": url,
                        "extraction_method": extraction_method,
                    }
    return meta_by_key


def _read_enriched_csv_dates() -> dict[str, dict[str, str]]:
    """Read enriched CSV → {source_key: {start_date, end_date}}."""
    stats: dict[str, dict[str, str]] = {}
    with ENRICHED_CSV.open(newline="", encoding="utf-8") as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            sk = (row.get("source_key") or "").strip()
            if not sk:
                continue
            d = (row.get("observation_date") or "").strip()
            if not d:
                continue
            if sk not in stats:
                stats[sk] = {"start_date": d, "end_date": d}
            else:
                if d < stats[sk]["start_date"]:
                    stats[sk]["start_date"] = d
                if d > stats[sk]["end_date"]:
                    stats[sk]["end_date"] = d
    return stats


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

_HEADERS = [
    ("source_name", "Source Name"),
    ("country", "Country"),
    ("url", "URL"),
    ("start_date", "Start Date"),
    ("end_date", "End Date"),
    ("data_format", "Data Format"),
    ("cadence", "Cadence"),
]


def build_rows() -> list[dict[str, str]]:
    """Assemble one row per source in the intersection of registry and enriched CSV."""
    print("Collecting SOURCE_META from fetcher modules ...")
    meta_by_key = _collect_source_meta()

    print(f"Reading enriched CSV: {ENRICHED_CSV} ...")
    csv_dates = _read_enriched_csv_dates()

    rows: list[dict[str, str]] = []
    for source_key in sorted(FETCHER_REGISTRY):
        if source_key not in csv_dates:
            print(f"  SKIP (not in enriched CSV): {source_key}")
            continue

        cfg = FETCHER_REGISTRY[source_key]
        meta = meta_by_key.get(source_key, {})

        url = meta.get("url") or cfg.homepage
        extraction_method = meta.get("extraction_method", [])
        data_format = _simplify_extraction_method(extraction_method)
        dates = csv_dates[source_key]

        rows.append(
            {
                "source_name": cfg.source_name,
                "country": cfg.country,
                "url": str(url),
                "start_date": dates["start_date"],
                "end_date": dates["end_date"],
                "data_format": data_format,
                "cadence": cfg.cadence,
            }
        )

    return rows


def write_xlsx(rows: list[dict[str, str]], output_path: Path) -> None:
    """Write rows to an Excel workbook with light formatting."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Source Catalog"

    # -- header row --------------------------------------------------------
    hdr_font = Font(bold=True)
    hdr_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for col_idx, (_key, label) in enumerate(_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")

    # -- data rows ---------------------------------------------------------
    link_font = Font(color="0563C1", underline="single")
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, (key, _label) in enumerate(_HEADERS, 1):
            value = row_data[key]
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if key == "url" and value.startswith("http"):
                cell.hyperlink = value
                cell.font = link_font

    # -- column widths -----------------------------------------------------
    for col_idx, (key, label) in enumerate(_HEADERS, 1):
        max_len = len(label)
        for row_idx in range(2, len(rows) + 2):
            val = str(ws.cell(row=row_idx, column=col_idx).value or "")
            max_len = max(max_len, len(val))
        # cap URL column at 60, others at 40
        cap = 60 if key == "url" else 40
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(
            max_len + 2, cap
        )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)


def main() -> None:
    rows = build_rows()
    print(f"Writing {len(rows)} sources to {OUTPUT_XLSX} ...")
    write_xlsx(rows, OUTPUT_XLSX)
    print("Done.")


if __name__ == "__main__":
    main()
